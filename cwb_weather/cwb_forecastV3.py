def weekForecast():
    from urllib import parse
    import requests as req
    from bs4 import BeautifulSoup as bs
    from fake_useragent import UserAgent
    from datetime import datetime, timedelta
    import re

    # ===== 產生取得最新預報用網址 =====
    # 預報網站的最新資訊html每10分鐘會變，格式基於目前日期時間(到10分鐘位)
    # import datetime物件用來抓取目前時間並動態產生query string加入url
    now = datetime.now()
    nowFormatted = f"{now.year}{now.month:02}{now.day:02}{now.hour:02}-{str(now.minute)[0]}"
    url = f"https://www.cwb.gov.tw/V8/C/W/County/MOD/wf7dayNC_NCSEI/ALL_Week.html?t={nowFormatted}"
    print(f"最新預報的網址是: {url}")
    # ===== 產生取得最新預報用網址 end =====

    # 設定隨機UserAgent
    ua = UserAgent()
    my_headers = {
        'User-Agent':ua.random
    }

    # 發出get request，確認連線狀況
    response = req.get(url, headers = my_headers)
    # print(f'網站狀態碼: {response.status_code}')
    # print(f'網站編碼　: {response.encoding}')
    # print(f'回覆標頭　: {response.headers}')

    # 建立美麗湯物件
    soup = bs(response.text, 'lxml')

    # ==== 建立變數說明 ==== 
    # dataByArea => 縣市天氣資料父標籤，共22筆
    # rawDate => 預報日期(表頭)資料標籤，共7筆
    # reDate => 以m/dd取出天氣資料日期標題的正規表達式
    # dayPartLabel => 存放資料到weeklyTempData的dayPart下時動態決定存放值的欄位
    # dateKeys => 存放資料到weeklyTempData時動態變換key用串列
    # weeklyTempData => 存放所有
    # ==== 建立變數說明 end ==== 

    dataByArea = soup.select('tbody')
    rawDate = soup.select('tr.table_top th:not(#County, #time)') # 本份1週預報每天的日期
    reDate = r'([\d]+\/[\d]{2})'
    weeklyTempData = {}

    # ==== 建立dayPart串列 ====
    dayPartLabel = []
    dayPartLabel.append(dataByArea[0].select('tr.day td[headers] span')[0].get_text()) # 白天
    dayPartLabel.append(dataByArea[0].select('tr.night td[headers] span')[0].get_text()) # 晚上
    # ==== 建立dayPart結構 end ====

    # ==== 建立weeklyTempData key (Part 1) ====
    weeklyTempData['cityCounty'] = []
    weeklyTempData['dayPart'] = []
    # ==== 建立weeklyTempData key (Part 1) end ====

    # ==== 建立weeklyTempData key (Part 2) ====
    dateKeys = []
    for rawDay in rawDate:
        day = re.search(reDate, rawDay.get_text())[0]
        weeklyTempData[day] = []
        dateKeys.append(day)
    # ==== 建立weeklyTempData key (Part 2) end ====

    # ==== for迴圈加資料進入weeklyTempData ====
    for county in dataByArea:
        cityName = county.select('th[headers] span')[0].get_text() # 縣市名稱，1個縣市1筆
        dayTempAll = county.select('tr.day td[headers] span.tem-C') # 白天溫度，1個縣市7筆
        nightTempAll = county.select('tr.night td[headers] span.tem-C') # 晚上溫度，1個縣市7筆
        
        # 因1個縣市會有白天以及晚上2種溫度，拆成2列來呈現 (白天、晚上各1列)
        for i in range(2):
            weeklyTempData['cityCounty'].append(cityName)
            weeklyTempData['dayPart'].append(dayPartLabel[i])
        
        # 取得白天、晚上溫度，去除多餘空格後，寫入各個對應日期的串列
        for i in range(len(dayTempAll)):
            dayTemp = re.sub(r'\u2002',r'',dayTempAll[i].get_text())
            nightTemp = re.sub(r'\u2002',r'',nightTempAll[i].get_text())
            weeklyTempData[dateKeys[i]].append(dayTemp)
            weeklyTempData[dateKeys[i]].append(nightTemp)
    
    return weeklyTempData
            

def outputCSV(weatherDict):
    import pandas as pd
    fileName = 'weeklyTempData.csv'
    weatherDict_df = pd.DataFrame(weatherDict)
    weatherDict_df.to_csv(fileName)
    print(f"最新1週縣市天氣預報資訊輸出至 {fileName}")

def outputExcel(weatherDict):
    from openpyxl import Workbook, load_workbook
    fileName = 'weeklyTempExcel.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = '一周天氣預報'

    title_row = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ' # 無法對應超過Z欄以上
    # ==== 建立表格結構 ====
    i = 0
    for key in weatherDict.keys():
        title_cell = f"{title_row[i]}1"
        ws[title_cell] = key
        j = 2
        for item in weatherDict[key]:
            current_cell = f"{title_row[i][0:1]}{j}"
            ws[current_cell] = item
            j += 1
        i += 1
    # ==== 創建excel表格標題列 end ====

    wb.save(fileName)
    print(f"最新1週縣市天氣預報資訊輸出至 {fileName}")

if __name__ == '__main__':
    weatherDict = weekForecast()
    outputCSV(weatherDict)
    outputExcel(weatherDict)